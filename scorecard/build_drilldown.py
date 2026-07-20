#!/usr/bin/env python3
"""Phase 5 drill-down workbook — renders the account/contact drill-down and the
owned-account neglect view into a presentable Excel for the PM / sales leaders.

READ-ONLY: only SELECTs from the Phase 5 functions (migrations/003). Meetings
are excluded from account views (they carry no company yet — v1 decision);
each rep's activities with no matched company appear as an explicit
'(no account matched)' row so totals reconcile instead of quietly shrinking.

Run:  python scorecard/build_drilldown.py
Output: reports/CA_account_drilldown_<today>.xlsx  (reports/ is gitignored —
named-account/rep data never goes into git history)
"""
import os
import re
import sys
from datetime import date

import psycopg2
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

_HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(_HERE, "..", "ingestion"))
from ingest import load_env, require

INK = "1F2937"
HEAD_FILL = PatternFill("solid", fgColor="123A32")
SUB_FILL = PatternFill("solid", fgColor="E8EFEC")
WARN_FILL = PatternFill("solid", fgColor="FDECEA")   # pale red: neglected top tier
NOTE_FONT = Font(name="Arial", size=9, italic=True, color="5C6B66")
H1 = Font(name="Arial", size=15, bold=True, color=INK)
HEAD_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
BODY = Font(name="Arial", size=10, color=INK)
BODY_B = Font(name="Arial", size=10, bold=True, color=INK)
thin = Side(style="thin", color="C9D2CD")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def style(cell, *, font=BODY, fmt=None, fill=None, align="right"):
    cell.font = font
    cell.border = BORDER
    if fmt:
        cell.number_format = fmt
    if fill:
        cell.fill = fill
    if align:
        cell.alignment = Alignment(horizontal=align)
    return cell


def header_row(ws, r, heads):
    for i, h in enumerate(heads, start=1):
        c = ws.cell(row=r, column=i, value=h)
        c.font = HEAD_FONT
        c.fill = HEAD_FILL
        c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def title(ws, text, note):
    ws.sheet_view.showGridLines = False
    ws["A1"] = text
    ws["A1"].font = H1
    ws["A2"] = note
    ws["A2"].font = NOTE_FONT


def put_rows(ws, r0, rows, text_cols):
    """Write value rows; text_cols are 1-based column indexes rendered left."""
    r = r0
    for row in rows:
        for c, v in enumerate(row, start=1):
            if c in text_cols:
                style(ws.cell(row=r, column=c, value=v if v is not None else ""),
                      font=BODY, align="left")
            elif isinstance(v, (int, float)):
                style(ws.cell(row=r, column=c, value=int(v)), fmt="#,##0")
            else:
                style(ws.cell(row=r, column=c, value=str(v) if v is not None else ""),
                      align="right")
        r += 1
    return r


def main():
    load_env()
    conn = psycopg2.connect(require("SUPABASE_DB_URL"), connect_timeout=20)
    conn.autocommit = True
    cur = conn.cursor()

    cur.execute("select min(activity_date), max(activity_date) from activity")
    d0, d1 = cur.fetchone()
    window = f"{d0} to {d1} (UTC), all data held"
    wb = Workbook()

    # ------------------------------------------------------- 1. team master
    ws = wb.active
    ws.title = "Team master"
    title(ws, "Rep x account drill-down - whole team",
          f"Window: {window}. Meetings excluded (not linkable to accounts yet). "
          "'(no account matched)' = activity the source logged without a company (~60%, mostly LinkedIn/calls) - shown, never hidden. "
          "Definitions: docs/ontology.md.")
    heads = ["CA", "Account", "Tier", "Owned by this rep", "Touchpoints", "People",
             "Auto email", "Manual email", "Calls", "LinkedIn", "Inbound", "First touch", "Last touch"]
    header_row(ws, 4, heads)
    cur.execute("""select ca_name, account_name, icp_tier,
                          case when owned_by_this_rep then 'yes' else '' end,
                          touchpoints, people_touched, auto_email, manual_email,
                          calls, linkedin, inbound_replies, first_touch, last_touch
                   from rep_account_drilldown_alltime
                   order by ca_name, touchpoints desc""")
    master = cur.fetchall()
    put_rows(ws, 5, master, text_cols={1, 2, 3, 4, 12, 13})
    ws.freeze_panes = "C5"
    for col, w in zip("ABCDEFGHIJKLM", [26, 34, 9, 12, 11, 8, 10, 11, 8, 9, 9, 11, 11]):
        ws.column_dimensions[col].width = w

    # -------------------------------------- 1b. owned-accounts-per-rep summary
    ws = wb.create_sheet("Owned per rep")
    title(ws, "Owned accounts per rep - coverage summary",
          f"Window: {window}. Owned = accounts assigned to the rep in HubSpot (target-account owner); "
          "a current snapshot, not window-dependent. Touched/coverage count activity in the window "
          "(excl. meetings and the ~60% no-company activity, so coverage is a floor). "
          "Sorted worst coverage first. Full account lists: 'Neglect - owned accounts'.")
    heads = ["CA", "Accounts owned", "Owned & touched", "Untouched", "Coverage %",
             "Tier 0/1 owned untouched"]
    header_row(ws, 4, heads)
    cur.execute("""
        select owner_name,
               count(*)                                         as owned,
               count(*) filter (where owner_touches > 0)        as touched,
               count(*) filter (where owner_touches = 0
                                  and icp_tier in ('Tier 0','Tier 1')) as t01_untouched
        from owned_account_coverage_alltime
        group by owner_name
        order by (count(*) filter (where owner_touches > 0))::float
                 / nullif(count(*),0) asc nulls last, owner_name
    """)
    rows = cur.fetchall()
    r = 5
    for name, owned, touched, t01 in rows:
        style(ws.cell(row=r, column=1, value=name), font=BODY_B, align="left")
        style(ws.cell(row=r, column=2, value=int(owned)), fmt="#,##0")
        style(ws.cell(row=r, column=3, value=int(touched)), fmt="#,##0")
        style(ws.cell(row=r, column=4, value=f"=B{r}-C{r}"), fmt="#,##0")
        style(ws.cell(row=r, column=5, value=f'=IF(B{r}=0,"n/a",C{r}/B{r})'), fmt="0.0%")
        c = style(ws.cell(row=r, column=6, value=int(t01)), fmt="#,##0")
        if t01:
            c.fill = WARN_FILL
        r += 1
    style(ws.cell(row=r, column=1, value="All CAs"), font=BODY_B, fill=SUB_FILL, align="left")
    for col in (2, 3, 4, 6):
        L = chr(64 + col)
        style(ws.cell(row=r, column=col, value=f"=SUM({L}5:{L}{r-1})"), font=BODY_B,
              fmt="#,##0", fill=SUB_FILL)
    style(ws.cell(row=r, column=5, value=f"=C{r}/B{r}"), font=BODY_B, fmt="0.0%", fill=SUB_FILL)
    ws.freeze_panes = "B5"
    for col, w in zip("ABCDEF", [26, 15, 16, 12, 12, 22]):
        ws.column_dimensions[col].width = w
    wb.move_sheet(ws, -(wb.index(ws)))   # make it the first tab

    # ------------------------------------------------- 2. neglect (owned)
    ws = wb.create_sheet("Neglect - owned accounts")
    title(ws, "Owned accounts incl. zeros - the neglect view",
          f"Window: {window}. One row per CA-owned account (HubSpot target-account owner). "
          "owner touches = by the owner; team touches = by ANY CA. Red rows = Tier 0/1 with zero touches from anyone. "
          "Touch counts exclude meetings and the ~60% of activity with no matched company, so a low number can undercount - treat as a floor.")
    heads = ["Owner", "Account", "Tier", "Vertical", "Owner touches", "Owner last touch",
             "Team touches", "Team last touch", "Reps on it"]
    header_row(ws, 4, heads)
    cur.execute("""select owner_name, account_name, icp_tier, vertical,
                          owner_touches, owner_last_touch, team_touches,
                          team_last_touch, team_reps
                   from owned_account_coverage_alltime
                   order by case when icp_tier='Tier 0' then 0
                                 when icp_tier='Tier 1' then 1
                                 when icp_tier='Tier 2' then 2
                                 when icp_tier='Tier 3' then 3
                                 when icp_tier='Tier 4' then 4 else 9 end,
                            team_touches, owner_name""")
    rows = cur.fetchall()
    r = put_rows(ws, 5, rows, text_cols={1, 2, 3, 4, 6, 8})
    # highlight neglected top-tier rows
    for i, row in enumerate(rows):
        if row[2] in ("Tier 0", "Tier 1") and (row[6] or 0) == 0:
            for c in range(1, len(heads) + 1):
                ws.cell(row=5 + i, column=c).fill = WARN_FILL
    ws.freeze_panes = "C5"
    for col, w in zip("ABCDEFGHI", [26, 34, 9, 22, 12, 13, 12, 13, 9]):
        ws.column_dimensions[col].width = w

    # ------------------------------------------------------- 3. contacts
    ws = wb.create_sheet("Contacts")
    title(ws, "Account -> contact drill-down - whole team",
          f"Window: {window}. One row per rep x account x person. Only activities with a matched person appear here "
          "(an account-level touch without a person shows on the account sheets, not this one). Meetings excluded.")
    heads = ["CA", "Account", "Tier", "Contact", "Job title", "Touchpoints",
             "Emails", "Calls", "LinkedIn", "Inbound", "Last touch"]
    header_row(ws, 4, heads)
    cur.execute("""select ca_name, account_name, icp_tier,
                          coalesce(contact_name, contact_email, contact_id) as who,
                          jobtitle, touchpoints, emails, calls, linkedin,
                          inbound_replies, last_touch
                   from account_contact_drilldown_alltime
                   order by ca_name, account_name, touchpoints desc""")
    put_rows(ws, 5, cur.fetchall(), text_cols={1, 2, 3, 4, 5, 11})
    ws.freeze_panes = "C5"
    for col, w in zip("ABCDEFGHIJK", [24, 30, 9, 26, 30, 11, 8, 8, 9, 8, 11]):
        ws.column_dimensions[col].width = w

    # ------------------------------------------------------ 4. per-rep tabs
    cur.execute("select ca_name from rep_scorecard_alltime order by total_counted desc")
    reps = [r[0] for r in cur.fetchall()]
    for rep in reps:
        tab = re.sub(r"[\\/*?:\[\]]", "", rep)[:31]
        ws = wb.create_sheet(tab)
        title(ws, f"{rep} - accounts worked",
              f"Window: {window}. Ranked by touchpoints. Meetings excluded (see rep scorecard for meetings). "
              "The '(no account matched)' row keeps the total honest.")
        heads = ["Account", "Tier", "Owned by rep", "Touchpoints", "People",
                 "Auto email", "Manual email", "Calls", "LinkedIn", "Inbound", "Last touch"]
        header_row(ws, 4, heads)
        cur.execute("""select account_name, icp_tier,
                              case when owned_by_this_rep then 'yes' else '' end,
                              touchpoints, people_touched, auto_email, manual_email,
                              calls, linkedin, inbound_replies, last_touch
                       from rep_account_drilldown_alltime
                       where ca_name = %s order by touchpoints desc""", (rep,))
        rows = cur.fetchall()
        r = put_rows(ws, 5, rows, text_cols={1, 2, 3, 11})
        style(ws.cell(row=r, column=1, value="Total"), font=BODY_B, fill=SUB_FILL, align="left")
        for c in range(2, len(heads) + 1):
            letter = chr(64 + c) if c <= 26 else None
            style(ws.cell(row=r, column=c,
                          value=f"=SUM({letter}5:{letter}{r-1})" if 4 <= c <= 10 else ""),
                  font=BODY_B, fmt="#,##0", fill=SUB_FILL)
        ws.freeze_panes = "B5"
        for col, w in zip("ABCDEFGHIJK", [34, 9, 12, 11, 8, 10, 11, 8, 9, 9, 11]):
            ws.column_dimensions[col].width = w

    conn.close()
    out = os.path.abspath(os.path.join(_HERE, "..", "reports",
                                       f"CA_account_drilldown_{date.today()}.xlsx"))
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)
    print("saved", out)


if __name__ == "__main__":
    main()
