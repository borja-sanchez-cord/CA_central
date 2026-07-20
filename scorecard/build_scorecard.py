#!/usr/bin/env python3
"""Phase 4 scorecard workbook — renders the rep_scorecard() Supabase function
into a presentable Excel for the PM / sales leaders.

READ-ONLY: this script only SELECTs from the Phase 4 views. The database
objects it reads are created by migrations/002_rep_scorecard.sql; the
definitions of every column are in docs/ontology.md.

Run:  python scorecard/build_scorecard.py
Output: reports/CA_rep_scorecard_<today>.xlsx  (reports/ is gitignored -
per-rep performance data never goes into git history)
"""
import os
import sys
from datetime import date, timedelta

import psycopg2
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

_HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(_HERE, "..", "ingestion"))
sys.path.insert(0, _HERE)
from ingest import load_env, require
import layout  # single source of truth for the column order/headers/formulas

INK = "1F2937"
HEAD_FILL = PatternFill("solid", fgColor="123A32")
SUB_FILL = PatternFill("solid", fgColor="E8EFEC")
NOTE_FONT = Font(name="Arial", size=9, italic=True, color="5C6B66")
H1 = Font(name="Arial", size=15, bold=True, color=INK)
HEAD_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
BODY = Font(name="Arial", size=10, color=INK)
BODY_B = Font(name="Arial", size=10, bold=True, color=INK)
thin = Side(style="thin", color="C9D2CD")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# Team-wide distinct accounts/contacts (NOT the per-rep sum: two reps can touch
# the same account). Keys map to the TEAM_DISTINCT_KEYS handled in the team row.
TEAM_DISTINCT = """
    select count(distinct company_id) filter (where company_id is not null),
           count(distinct contact_id) filter (where contact_id is not null)
    from activity_flat
    where counts and ca_id is not null and activity_date between %s and %s
"""


def _style(cell, *, font, fmt=None, fill=None, align="right"):
    cell.font = font
    cell.border = BORDER
    if fmt:
        cell.number_format = fmt
    if fill:
        cell.fill = fill
    if align:
        cell.alignment = Alignment(horizontal=align)
    return cell


def _formula(key, r):
    """Spreadsheet formula for a computed column, referenced BY KEY (so it
    can't point at the wrong letter if the layout changes)."""
    _k, _h, kind, refs = next(c for c in layout.COLUMNS if c[0] == key)
    a, b = (layout.column_letter(refs[0]), layout.column_letter(refs[1]))
    if kind == "sum":
        return f"={a}{r}+{b}{r}"
    if kind == "ratio":
        return f'=IF({b}{r}=0,"n/a",{a}{r}/{b}{r})'
    raise ValueError(f"{key} is not a computed column")


def sheet(wb, title, window_label, rows, team_accounts, team_contacts, first=False):
    ws = wb.active if first else wb.create_sheet()
    ws.title = title
    ws.sheet_view.showGridLines = False
    ws["A1"] = f"CA rep scorecard - {window_label}"
    ws["A1"].font = H1
    ws["A2"] = ("Source: rep_scorecard() view in Supabase (Phase 4), read-only over activity_flat. "
                "Definitions: docs/ontology.md. Meetings are BOOKED unless the held column says otherwise.")
    ws["A2"].font = NOTE_FONT

    read_keys = layout.db_keys()          # order rows arrive from SQL
    row_pos = {k: i for i, k in enumerate(read_keys)}
    team_distinct = {"accounts_touched": team_accounts, "contacts_touched": team_contacts}

    # header row
    r = 4
    for col, header in enumerate(layout.headers(), start=1):
        _style(ws.cell(row=r, column=col, value=header), font=HEAD_FONT, fill=HEAD_FILL,
               align=None)
        ws.cell(row=r, column=col).alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True)
    r += 1
    r0 = r

    # per-rep rows
    for row in rows:
        for col, (key, _h, kind, _refs) in enumerate(layout.COLUMNS, start=1):
            if kind == "text":
                _style(ws.cell(row=r, column=col, value=row[row_pos[key]]),
                       font=BODY_B, align="left")
            elif kind == "db":
                _style(ws.cell(row=r, column=col, value=int(row[row_pos[key]])),
                       font=BODY, fmt="#,##0")
            elif kind == "sum":
                _style(ws.cell(row=r, column=col, value=_formula(key, r)),
                       font=BODY, fmt="#,##0")
            elif kind == "ratio":
                _style(ws.cell(row=r, column=col, value=_formula(key, r)),
                       font=BODY, fmt="0.0%")
        r += 1
    r_team = r

    # team totals row
    for col, (key, _h, kind, _refs) in enumerate(layout.COLUMNS, start=1):
        letter = layout.column_letter(key)
        if kind == "text":
            _style(ws.cell(row=r, column=col, value=f"All {len(rows)} CAs"),
                   font=BODY_B, fill=SUB_FILL, align="left")
        elif key in layout.TEAM_DISTINCT_KEYS:
            # true distinct across reps, not the column sum
            _style(ws.cell(row=r, column=col, value=int(team_distinct[key])),
                   font=BODY_B, fmt="#,##0", fill=SUB_FILL)
        elif kind == "ratio":
            a, b = layout.column_letter(_refs[0]), layout.column_letter(_refs[1])
            _style(ws.cell(row=r, column=col, value=f"={a}{r}/{b}{r}"),
                   font=BODY_B, fmt="0.0%", fill=SUB_FILL)
        else:  # db + sum both roll up as a column sum
            _style(ws.cell(row=r, column=col, value=f"=SUM({letter}{r0}:{letter}{r-1})"),
                   font=BODY_B, fmt="#,##0", fill=SUB_FILL)

    ws.cell(row=r_team + 2, column=1, value=(
        "Coverage % = of the accounts THIS rep owns (HubSpot target-account owner), the share they touched "
        "in the window. Accounts/Contacts touched skip the ~60% of activities with no matched company - "
        "they understate breadth; watch the trend, not the level. Team accounts/contacts are true distincts, "
        "not the per-rep sum. Pursuits/Conversations are subsets of Dials; Total counts every activity once.")
    ).font = NOTE_FONT
    ws.freeze_panes = "B5"
    ws.column_dimensions["A"].width = 28
    for col in range(2, len(layout.COLUMNS) + 1):
        ws.column_dimensions[layout.column_letter(layout.COLUMNS[col - 1][0])].width = 11
    return ws


def main():
    load_env()
    conn = psycopg2.connect(require("SUPABASE_DB_URL"), connect_timeout=20)
    conn.autocommit = True
    cur = conn.cursor()

    today = date.today()
    # "All time" starts at the earliest day we actually hold, read from the
    # data — no hardcoded go-live date to drift.
    cur.execute("select min(activity_date) from activity")
    data_start = cur.fetchone()[0] or date(2026, 7, 6)
    windows = [
        ("Last 7 days", today - timedelta(days=7), today - timedelta(days=1)),
        ("All time", data_start, today - timedelta(days=1)),
    ]

    select_sql = layout.select_sql()  # projection built from the column spec

    wb = Workbook()
    first = True
    for label, start, end in windows:
        cur.execute(select_sql, (start, end))
        rows = cur.fetchall()
        cur.execute(TEAM_DISTINCT, (start, end))
        team_accounts, team_contacts = cur.fetchone()
        sheet(wb, label, f"{start} to {end} (UTC)", rows,
              team_accounts or 0, team_contacts or 0, first=first)
        first = False
    conn.close()

    out = os.path.join(_HERE, "..", "reports", f"CA_rep_scorecard_{today}.xlsx")
    out = os.path.abspath(out)
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)
    print("saved", out)


if __name__ == "__main__":
    main()
